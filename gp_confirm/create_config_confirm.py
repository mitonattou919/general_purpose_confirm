# coding: utf-8

import openpyxl
import sys
import os

# Get target Excel file name.
args = sys.argv
my_wb = openpyxl.load_workbook(args[1], data_only=True)

# Define file name.
inventory_file = 'hosts'
var_dir_host   = 'host_vars'
var_dir_group  = 'group_vars'

max_target = 16                # Max number of target host.

class Color:
    BLACK = '\033[30m'
    RED   = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    RESET = '\033[0m'

# [ START write_file]
# Write strings to file.
# Args:
#   file_name: specify target file name.
#   text: text strings to write.
# Returns:
#   none
def write_file(file_name, text):
    with open(file_name, mode='a') as f:
        f.write(text)
# [ END write_file]

# [ START remove_file]
# Remove file.
# Args:
#   file_name: specify target file name.
# Returns:
#   none
def remove_file(file_name):
    if os.path.exists(file_name):
        os.remove(file_name)
# [ END remove_file]


# [ START create_inventory]
# create inventory file and initilize files.
# Args:
#   none
# Returns:
#   none
def create_inventory():
    env_values = []
    host_list = []
    my_ws = my_wb['env']

    for col in my_ws.iter_cols(min_row=5, min_col=1, max_row=10):
        env_values.append([cell.value for cell in col])

    ## File initialization.
    remove_file(inventory_file)

    ## Create inventory file
    for value in env_values:
        if value[0] != None:
            text = ('[' + str(value[2]) + ']\n' 
                    + str(value[1]) + ' '
                    + env_values[0][3] + '=' + str(value[3]) + ' '
                    + env_values[0][4] + '=' + str(value[4]) + ' '
                    + env_values[0][5] + '=' + str(value[5]) + '\n\n')
            write_file(inventory_file, text)
            host_list.append(value[1])

    # Count number of target host.
    num_target = len(host_list)

    print('\nTarget host:')

    # Initialize variable file.
    for i in range(num_target):
        print(host_list[i])
        remove_file(var_dir_host + '/' + str(host_list[i]) + '.yml')
# [ END create_inventory]


# [ START create_vars]
# Create variable file.
# Args:
#   sheet_name: specify target sheet name.
#   parm_name: specify target parameter name.
# Returns:
#   none
def create_vars(sheet_name, parm_name):
    ws_values = []
    my_ws = my_wb[sheet_name]
    text = ''
    last_texts = {}

    for row in my_ws.iter_rows(min_row=1, min_col=1):
        ws_values.append([cell.value for cell in row])

    for value in ws_values:
        if value[0] == parm_name:
            text = '\n' + str(value[1]) + '\n' + str(value[0]) + ':\n'
            for i in range(max_target + 10):
                if i >= 10 and value[i] is not None:
                    write_file(var_dir_host + '/' + str(ws_values[1][i]) + '.yml', text)
                    last_texts[ws_values[1][i]] = text
        else:
            for i in range(max_target + 10):
                if i >= 10 and value[i] is not None and value[0] != 'product':
                    if ws_values[1][2] == 'parameter':
                        text = (
                            ' - { '
                            + str(ws_values[1][1]) + ': \'' + str(value[1]) + '\''
                            + ' }\n')
                    elif ws_values[1][3] == 'parameter':
                        text = (
                            ' - { '
                            + str(ws_values[1][1]) + ': \'' + str(value[1]) + '\', '
                            + str(ws_values[1][2]) + ': \'' + str(value[2]) + '\''
                            + ' }\n')

                    if text != last_texts[ws_values[1][i]]:
                        write_file(var_dir_host + '/' + str(ws_values[1][i]) + '.yml', text)
                        last_texts[ws_values[1][i]] = text
# [ END create_vars]


# [ START main]
def main():

    # Confirm exectuion environment is correct.
    ws_env = my_wb['env']
    env_type = ws_env.cell(2, 2).value

    question_msg = ('You execute this script in ' 
                    + Color.RED + env_type + Color.RESET
                    + ' environment.\nAre you sure? [y/N]: ')
    user_answer = input(question_msg)

    if user_answer != 'y':
        exit()

    # Inventory
    create_inventory()

    # Files
    create_vars('files', 'general_purpose_copy_info')

    # Commands
    create_vars('cmd', 'general_purpose_cmd_info')

    print('\nInventory file and variables files has created.')

# [ END main]

if __name__ == "__main__":
    main()
